"""
Vues Django pour TRANS-GEST.
Système de Gestion des Matériels de Transmission
"""
import io
import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import HttpResponse

from .models import Magasin, Detenteur, Materiel, Document, Mouvement, \
    CategorieMaterielLogistique, LigneInventaireLogistique
from .forms import MagasinForm, DetenteurForm, MaterielForm, DocumentForm, MouvementForm


# ===========================================================================
# AUTHENTICATION
# ===========================================================================

def login_view(request):
    """Vue de connexion."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Bienvenue, {user.username} !')
            return redirect('welcome')
        else:
            messages.error(request, 'Identifiants invalides.')
    else:
        form = AuthenticationForm()
    return render(request, 'auth/login.html', {'form': form})


def logout_view(request):
    """Vue de déconnexion."""
    logout(request)
    messages.info(request, 'Vous avez été déconnecté.')
    return redirect('login')


# Admin-only User Management
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test

def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_admin)
def user_list(request):
    """Vue de gestion des utilisateurs (Admin uniquement)."""
    users = User.objects.all().order_by('-date_joined')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')
            email = request.POST.get('email', '').strip()
            is_staff = request.POST.get('is_staff') == 'on'
            
            if User.objects.filter(username=username).exists():
                messages.error(request, "Cet identifiant existe déjà.")
            else:
                user = User.objects.create_user(username=username, email=email, password=password)
                user.is_staff = is_staff
                user.save()
                messages.success(request, f"Utilisateur {username} ajouté avec succès.")
        
        elif action == 'delete':
            user_id = request.POST.get('user_id')
            if int(user_id) == request.user.id:
                messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
            else:
                user = get_object_or_404(User, id=user_id)
                username = user.username
                user.delete()
                messages.success(request, f"Utilisateur {username} supprimé.")
        
        return redirect('user_list')
        
    return render(request, 'auth/user_list.html', {'users': users})


# ===========================================================================
# DASHBOARD
# ===========================================================================

@login_required
def welcome_view(request):
    """Vue d'accueil unifiée avec statistiques et inventaire complet."""
    stats = {
        'total_materiels': Materiel.objects.count(),
        'en_service': Materiel.objects.filter(etat='service').count(),
        'en_attente': Materiel.objects.filter(etat='attente').count(),
        'hors_service': Materiel.objects.filter(etat='hors_service').count(),
        'approvisionnement': Materiel.objects.filter(etat='approvisionnement').count(),
        'total_magasins': Magasin.objects.count(),
        'total_detenteurs': Detenteur.objects.count(),
        'total_mouvements': Mouvement.objects.count(),
        'total_documents': Document.objects.count(),
    }

    # Tous les matériels pour la grille
    materiels = Materiel.objects.select_related('magasin', 'detenteur').all()

    # Derniers mouvements
    derniers_mouvements = Mouvement.objects.select_related(
        'materiel', 'document'
    ).order_by('-date')[:5]

    # Données pour les graphiques (par type de matériel)
    types_data = (
        Materiel.objects.values('type_materiel')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Données pour le graphique par état
    etats_data = (
        Materiel.objects.values('etat')
        .annotate(count=Count('id'))
        .order_by('etat')
    )

    context = {
        'stats': stats,
        'materiels': materiels,
        'derniers_mouvements': derniers_mouvements,
        'types_data': list(types_data),
        'etats_data': list(etats_data),
    }
    return render(request, 'welcome.html', context)


@login_required
def dashboard(request):
    """Redirige vers l'accueil unifié."""
    return redirect('welcome')


@login_required
def situation_view(request):
    """
    Vue générant le rapport 'Situation des Matériels' (Situation Globale).
    Agrège les données par type et désignation pour un affichage tabulaire.
    """
    from .models import Materiel
    
    # Récupérer tous les matériels
    materiels = Materiel.objects.select_related('magasin', 'detenteur').all()
    
    # Organiser les données par type de matériel
    situation_data = []
    
    for type_code, type_label in Materiel.TYPE_CHOICES:
        type_materiels = materiels.filter(type_materiel=type_code)
        
        if not type_materiels.exists():
            continue
            
        # Grouper par désignation
        designations_map = {}
        for m in type_materiels:
            designation_name = m.designation
            if designation_name not in designations_map:
                designations_map[designation_name] = {
                    'designation': designation_name,
                    'total': 0,
                    'magasin_counts': {},  # {magasin_name: count}
                    'detenteur_counts': {}, # {detenteur_name: count}
                    'observations': []
                }
            
            entry = designations_map[designation_name]
            entry['total'] = entry.get('total', 0) + 1

            
            if m.magasin:
                name = m.magasin.nom
                entry['magasin_counts'][name] = entry['magasin_counts'].get(name, 0) + 1
            
            if m.detenteur:
                name = m.detenteur.nom
                entry['detenteur_counts'][name] = entry['detenteur_counts'].get(name, 0) + 1
                
        situation_data.append({
            'type_label': type_label,
            'items': sorted(designations_map.values(), key=lambda x: x['designation'])
        })
    
    return render(request, 'situation_report.html', {
        'situation_data': situation_data,
        'title': 'Situation Globale des Matériels'
    })



# ===========================================================================
# SEARCH
# ===========================================================================

@login_required
def search_view(request):
    """Recherche globale."""
    query = request.GET.get('q', '').strip()
    results = {
        'materiels': [],
        'magasins': [],
        'detenteurs': [],
        'documents': [],
    }
    if query:
        results['materiels'] = Materiel.objects.filter(
            Q(designation__icontains=query) |
            Q(numero_serie__icontains=query) |
            Q(type_materiel__icontains=query)
        ).select_related('magasin', 'detenteur')[:20]

        results['magasins'] = Magasin.objects.filter(
            Q(nom__icontains=query) | Q(localisation__icontains=query)
        )[:10]

        results['detenteurs'] = Detenteur.objects.filter(
            Q(nom__icontains=query) | Q(fonction__icontains=query)
        )[:10]

        results['documents'] = Document.objects.filter(
            Q(description__icontains=query) | Q(observations__icontains=query)
        )[:10]

    return render(request, 'search_results.html', {'query': query, 'results': results})


# ===========================================================================
# MATERIEL CRUD
# ===========================================================================

@login_required
def materiel_list(request):
    """Liste des matériels avec filtres et pagination."""
    materiels = Materiel.objects.select_related('magasin', 'detenteur').all()

    # Filtres
    etat = request.GET.get('etat')
    type_m = request.GET.get('type')
    magasin_id = request.GET.get('magasin')
    q = request.GET.get('q', '').strip()

    if etat:
        materiels = materiels.filter(etat=etat)
    if type_m:
        materiels = materiels.filter(type_materiel=type_m)
    if magasin_id:
        materiels = materiels.filter(magasin_id=magasin_id)
    if q:
        materiels = materiels.filter(
            Q(designation__icontains=q) | Q(numero_serie__icontains=q)
        )

    paginator = Paginator(materiels, 20)
    page = request.GET.get('page')
    materiels_page = paginator.get_page(page)

    context = {
        'materiels': materiels_page,
        'magasins': Magasin.objects.all(),
        'etat_choices': Materiel.ETAT_CHOICES,
        'type_choices': Materiel.TYPE_CHOICES,
        'current_etat': etat,
        'current_type': type_m,
        'current_magasin': magasin_id,
        'current_q': q,
    }
    return render(request, 'materiel_list.html', context)


@login_required
def materiel_detail(request, pk):
    """Détail d'un matériel."""
    materiel = get_object_or_404(
        Materiel.objects.select_related('magasin', 'detenteur'), pk=pk
    )
    mouvements = materiel.mouvements.select_related('document').order_by('-date')[:10]
    return render(request, 'materiel_detail.html', {'materiel': materiel, 'mouvements': mouvements})


@login_required
def materiel_create(request):
    """Création d'un matériel."""
    if request.method == 'POST':
        form = MaterielForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Matériel créé avec succès.')
            return redirect('materiel_list')
    else:
        form = MaterielForm()
    return render(request, 'materiel_form.html', {'form': form, 'title': 'Nouveau Matériel'})


@login_required
def materiel_update(request, pk):
    """Modification d'un matériel."""
    materiel = get_object_or_404(Materiel, pk=pk)
    if request.method == 'POST':
        form = MaterielForm(request.POST, request.FILES, instance=materiel)
        if form.is_valid():
            form.save()
            messages.success(request, 'Matériel modifié avec succès.')
            return redirect('materiel_detail', pk=pk)
    else:
        form = MaterielForm(instance=materiel)
    return render(request, 'materiel_form.html', {'form': form, 'title': 'Modifier Matériel', 'object': materiel})


@login_required
def materiel_delete(request, pk):
    """Suppression d'un matériel."""
    materiel = get_object_or_404(Materiel, pk=pk)
    if request.method == 'POST':
        materiel.delete()
        messages.success(request, 'Matériel supprimé avec succès.')
        return redirect('materiel_list')
    return render(request, 'confirm_delete.html', {'object': materiel, 'type': 'matériel'})


# ===========================================================================
# MAGASIN CRUD
# ===========================================================================

@login_required
def magasin_list(request):
    """Liste des magasins."""
    magasins = Magasin.objects.annotate(nb_materiels=Count('materiels')).all()
    paginator = Paginator(magasins, 20)
    page = request.GET.get('page')
    return render(request, 'magasin_list.html', {'magasins': paginator.get_page(page)})


@login_required
def magasin_create(request):
    """Création d'un magasin."""
    if request.method == 'POST':
        form = MagasinForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Magasin créé avec succès.')
            return redirect('magasin_list')
    else:
        form = MagasinForm()
    return render(request, 'magasin_form.html', {'form': form, 'title': 'Nouveau Magasin'})


@login_required
def magasin_update(request, pk):
    """Modification d'un magasin."""
    magasin = get_object_or_404(Magasin, pk=pk)
    if request.method == 'POST':
        form = MagasinForm(request.POST, instance=magasin)
        if form.is_valid():
            form.save()
            messages.success(request, 'Magasin modifié avec succès.')
            return redirect('magasin_list')
    else:
        form = MagasinForm(instance=magasin)
    return render(request, 'magasin_form.html', {'form': form, 'title': 'Modifier Magasin', 'object': magasin})


@login_required
def magasin_delete(request, pk):
    """Suppression d'un magasin."""
    magasin = get_object_or_404(Magasin, pk=pk)
    if request.method == 'POST':
        magasin.delete()
        messages.success(request, 'Magasin supprimé avec succès.')
        return redirect('magasin_list')
    return render(request, 'confirm_delete.html', {'object': magasin, 'type': 'magasin'})


# ===========================================================================
# DETENTEUR CRUD
# ===========================================================================

@login_required
def detenteur_list(request):
    """Liste des détenteurs."""
    detenteurs = Detenteur.objects.annotate(nb_materiels=Count('materiels')).all()
    paginator = Paginator(detenteurs, 20)
    page = request.GET.get('page')
    return render(request, 'detenteur_list.html', {'detenteurs': paginator.get_page(page)})


@login_required
def detenteur_create(request):
    """Création d'un détenteur."""
    if request.method == 'POST':
        form = DetenteurForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Détenteur créé avec succès.')
            return redirect('detenteur_list')
    else:
        form = DetenteurForm()
    return render(request, 'detenteur_form.html', {'form': form, 'title': 'Nouveau Détenteur'})


@login_required
def detenteur_update(request, pk):
    """Modification d'un détenteur."""
    detenteur = get_object_or_404(Detenteur, pk=pk)
    if request.method == 'POST':
        form = DetenteurForm(request.POST, request.FILES, instance=detenteur)
        if form.is_valid():
            form.save()
            messages.success(request, 'Détenteur modifié avec succès.')
            return redirect('detenteur_list')
    else:
        form = DetenteurForm(instance=detenteur)
    return render(request, 'detenteur_form.html', {'form': form, 'title': 'Modifier Détenteur', 'object': detenteur})


@login_required
def detenteur_delete(request, pk):
    """Suppression d'un détenteur."""
    detenteur = get_object_or_404(Detenteur, pk=pk)
    if request.method == 'POST':
        detenteur.delete()
        messages.success(request, 'Détenteur supprimé avec succès.')
        return redirect('detenteur_list')
    return render(request, 'confirm_delete.html', {'object': detenteur, 'type': 'détenteur'})


# ===========================================================================
# DOCUMENT CRUD
# ===========================================================================

@login_required
def document_list(request):
    """Liste des documents."""
    documents = Document.objects.annotate(nb_mouvements=Count('mouvements')).all()
    type_filter = request.GET.get('type')
    if type_filter:
        documents = documents.filter(type=type_filter)
    paginator = Paginator(documents, 20)
    page = request.GET.get('page')
    return render(request, 'document_list.html', {
        'documents': paginator.get_page(page),
        'type_choices': Document.TYPE_CHOICES,
        'current_type': type_filter,
    })


@login_required
def document_create(request):
    """Création d'un document."""
    if request.method == 'POST':
        form = DocumentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Document créé avec succès.')
            return redirect('document_list')
    else:
        form = DocumentForm()
    return render(request, 'document_form.html', {'form': form, 'title': 'Nouveau Document'})


@login_required
def document_update(request, pk):
    """Modification d'un document."""
    document = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        form = DocumentForm(request.POST, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, 'Document modifié avec succès.')
            return redirect('document_list')
    else:
        form = DocumentForm(instance=document)
    return render(request, 'document_form.html', {'form': form, 'title': 'Modifier Document', 'object': document})


@login_required
def document_delete(request, pk):
    """Suppression d'un document."""
    document = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        document.delete()
        messages.success(request, 'Document supprimé avec succès.')
        return redirect('document_list')
    return render(request, 'confirm_delete.html', {'object': document, 'type': 'document'})


# ===========================================================================
# MOUVEMENT CRUD
# ===========================================================================

@login_required
def mouvement_list(request):
    """Liste des mouvements."""
    mouvements = Mouvement.objects.select_related('materiel', 'document').all()
    type_filter = request.GET.get('type')
    if type_filter:
        mouvements = mouvements.filter(type=type_filter)
    paginator = Paginator(mouvements, 20)
    page = request.GET.get('page')
    return render(request, 'mouvement_list.html', {
        'mouvements': paginator.get_page(page),
        'type_choices': Mouvement.TYPE_CHOICES,
        'current_type': type_filter,
    })


@login_required
def mouvement_create(request):
    """Création d'un mouvement."""
    if request.method == 'POST':
        form = MouvementForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mouvement enregistré avec succès.')
            return redirect('mouvement_list')
    else:
        form = MouvementForm()
    return render(request, 'mouvement_form.html', {'form': form, 'title': 'Nouveau Mouvement'})


@login_required
def mouvement_update(request, pk):
    """Modification d'un mouvement."""
    mouvement = get_object_or_404(Mouvement, pk=pk)
    if request.method == 'POST':
        form = MouvementForm(request.POST, instance=mouvement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mouvement modifié avec succès.')
            return redirect('mouvement_list')
    else:
        form = MouvementForm(instance=mouvement)
    return render(request, 'mouvement_form.html', {'form': form, 'title': 'Modifier Mouvement', 'object': mouvement})


@login_required
def mouvement_delete(request, pk):
    """Suppression d'un mouvement."""
    mouvement = get_object_or_404(Mouvement, pk=pk)
    if request.method == 'POST':
        mouvement.delete()
        messages.success(request, 'Mouvement supprimé avec succès.')
        return redirect('mouvement_list')
    return render(request, 'confirm_delete.html', {'object': mouvement, 'type': 'mouvement'})


# ===========================================================================
# EXPORT FUNCTIONS
# ===========================================================================

@login_required
def export_materiels_excel(request):
    """Exporter les matériels en Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Matériels'

    # En-tête style
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['ID', 'Désignation', 'Type', 'N° Série', 'État', 'Magasin', 'Détenteur', 'Date Sortie', 'Date Retour']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    materiels = Materiel.objects.select_related('magasin', 'detenteur').all()
    for row, m in enumerate(materiels, 2):
        ws.cell(row=row, column=1, value=m.id).border = thin_border
        ws.cell(row=row, column=2, value=m.designation).border = thin_border
        ws.cell(row=row, column=3, value=m.get_type_materiel_display()).border = thin_border
        ws.cell(row=row, column=4, value=m.numero_serie).border = thin_border
        ws.cell(row=row, column=5, value=m.get_etat_display()).border = thin_border
        ws.cell(row=row, column=6, value=m.magasin.nom).border = thin_border
        ws.cell(row=row, column=7, value=str(m.detenteur) if m.detenteur else '—').border = thin_border
        ws.cell(row=row, column=8, value=m.date_sortie.strftime('%d/%m/%Y') if m.date_sortie else '—').border = thin_border
        ws.cell(row=row, column=9, value=m.date_retour.strftime('%d/%m/%Y') if m.date_retour else '—').border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_length, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=materiels_{datetime.date.today().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
def export_materiels_csv(request):
    """Exporter les matériels en CSV."""
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename=materiels_{datetime.date.today().strftime("%Y%m%d")}.csv'
    response.write('\ufeff')  # BOM for Excel UTF-8

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['ID', 'Désignation', 'Type', 'N° Série', 'État', 'Magasin', 'Détenteur', 'Date Sortie', 'Date Retour'])

    for m in Materiel.objects.select_related('magasin', 'detenteur').all():
        writer.writerow([
            m.id, m.designation, m.get_type_materiel_display(),
            m.numero_serie, m.get_etat_display(), m.magasin.nom,
            str(m.detenteur) if m.detenteur else '—',
            m.date_sortie.strftime('%d/%m/%Y') if m.date_sortie else '—',
            m.date_retour.strftime('%d/%m/%Y') if m.date_retour else '—',
        ])
    return response


@login_required
def export_materiels_pdf(request):
    """Exporter les matériels en PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=materiels_{datetime.date.today().strftime("%Y%m%d")}.pdf'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#1E40AF'))
    elements.append(Paragraph('TRANS-GEST — Liste des Matériels', title_style))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f'Date: {datetime.date.today().strftime("%d/%m/%Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    # Table data
    data = [['#', 'Désignation', 'Type', 'N° Série', 'État', 'Magasin', 'Détenteur']]
    for m in Materiel.objects.select_related('magasin', 'detenteur').all():
        data.append([
            str(m.id), m.designation[:30], m.get_type_materiel_display(),
            m.numero_serie, m.get_etat_display(), m.magasin.nom,
            str(m.detenteur)[:20] if m.detenteur else '—',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4FF')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


# ===========================================================================
# INVENTAIRE LOGISTIQUE
# ===========================================================================

@login_required
def inventaire_logistique(request):
    """
    Vue du tableau logistique militaire.
    Affiche le matériel groupé par catégorie avec position et répartition.
    """
    categories = CategorieMaterielLogistique.objects.prefetch_related('lignes').order_by('ordre')

    # Calculer les totaux par catégorie
    data = []
    for cat in categories:
        lignes = list(cat.lignes.all())
        totaux = {
            'qte': sum(l.qte for l in lignes),
            'svc': sum(l.svc for l in lignes),
            'mag': sum(l.mag for l in lignes),
            'rep': sum(l.rep for l in lignes),
            'gmi': sum(l.gmi for l in lignes),
            'gs':  sum(l.gs  for l in lignes),
            'gcs': sum(l.gcs for l in lignes),
        }
        data.append({
            'categorie': cat,
            'lignes':    lignes,
            'totaux':    totaux,
        })

    return render(request, 'inventaire_logistique.html', {
        'data':  data,
        'title': 'Inventaire Logistique',
    })
